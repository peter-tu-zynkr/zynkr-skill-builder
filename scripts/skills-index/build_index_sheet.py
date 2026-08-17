#!/usr/bin/env python3
"""Build the Zynkr Skills Index workbook from inventory + extracted metadata.

Layout deliberately mirrors "[3.3] PM Flow x Skill Portfolio Assessment" (Sheet 15uxtdPh...):
  - Overview tab = 2-column label/value (Purpose / Sources / How to read / Legend / Headline / Verdict / Related)
  - Skill Index tab = one skill per row, parent skills bundling their children as
    child rows that carry a repeating No. and a leading  glyph in the parent column
  - trailing tabs are the flat evidence pivots

Brand tokens copied from 6.0 tech/zynkr-website-fe/styles.css :root (sync 2026-08-17).
Orange (#F26B1F) is the decision colour - used once, on the WIP status flag.
"""
import json, os, collections, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
J = os.path.join(HERE, 'data')
OUT = os.path.join(HERE, 'Zynkr-Skills-Index.xlsx')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PAPER, PAPER_RAIS = 'FBF7F0', 'FDFAF5'
INK, INK_SOFT, INK_MID = '0F0F0E', '2B2B28', '3D3A36'
SAGE_DEEP, SAGE_SOFT = '5FA48A', 'DCEEE6'
ORANGE, MUTE = 'F26B1F', '6F6B62'
FONT_ZH, FONT_MONO = 'Noto Sans TC', 'JetBrains Mono'

thin = Side(style='thin', color='E3DCD0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CATS = {
    'strategy':         ('0', 'Strategy & Leadership', '策略與領導'),
    'brand-marketing':  ('1', 'Brand & Marketing', '品牌與行銷'),
    'sales-consultant': ('2', 'Sales & Consultant', '銷售與顧問'),
    'operations':       ('3', 'Operations', '營運'),
    'training':         ('4', 'Training', '培訓'),
    'product':          ('5', 'Product', '產品'),
    'engineer':         ('6', 'Engineer', '工程'),
    'people-talent':    ('7', 'People & Talent', '人才與招募'),
    'finance-admin':    ('8', 'Finance & Admin', '財務與行政'),
    'legal':            ('9', 'Legal', '法務'),
}
KIND = {'orchestrator': 'Orchestrator', 'skill': 'Skill', 'subagent': 'Sub-skill'}
SIDE = {
    'read-only': 'Read-only',
    'creates-local-files': 'Creates local files',
    'creates-drafts': 'Creates drafts (never sends)',
    'writes-database': 'Writes database',
    'publishes-live': 'Publishes live',
    'mixed': 'Mixed',
}
KST = {
    'google-doc': 'Google Doc', 'google-sheet': 'Google Sheet', 'google-slides': 'Google Slides',
    'drive-folder': 'Drive folder', 'gmail': 'Gmail', 'supabase-kb': 'Supabase KB',
    'supabase-table': 'Supabase table', 'repo-file': 'Repo file', 'local-file': 'Local file',
    'github': 'GitHub', 'web': 'Web', 'calendar': 'Calendar', 'other': 'Other',
}


def cat_label(slug):
    n, en, zh = CATS.get(slug, ('9', slug, slug))
    return f'{n}. {en} {zh}'


import re

EXT_MAP = {
    'google drive': 'Google Drive', 'google drive/docs': 'Google Drive', 'google docs': 'Google Docs',
    'google sheets': 'Google Sheets', 'google slides': 'Google Slides', 'google calendar': 'Google Calendar',
    'google calendar appointment scheduling': 'Google Calendar', 'google chat': 'Google Chat',
    'google forms': 'Google Forms', 'google meet': 'Google Meet',
    'google search': 'Google Search', 'google search (serp)': 'Google Search',
    'gmail': 'Gmail', 'gmail (draft only)': 'Gmail', 'gmail (optional draft)': 'Gmail',
    'github': 'GitHub', 'accupass': 'Accupass 活動通', 'accupass 活動通': 'Accupass 活動通',
    'accupass / 活動通': 'Accupass 活動通', 'kit': 'Kit', 'notion': 'Notion', 'lucid': 'Lucid',
    'buffer': 'Buffer', 'threads': 'Threads', 'instagram': 'Instagram', 'facebook': 'Facebook',
    'line': 'LINE', 'supabase': 'Supabase', 'websearch': 'Web search', 'web search': 'Web search',
    'skills.sh': 'skills.sh', 'npm / npx skills cli': 'skills.sh',
    'ubersuggest': 'Ubersuggest', 'answerthepublic': 'AnswerThePublic',
    'claude.ai scheduled trigger': 'claude.ai routines', 'claude.ai routines': 'claude.ai routines',
    'local whisper': 'Whisper (local)', 'chrome/chromium': 'Chrome (CDP)',
}


def norm_ext(s):
    """Collapse free-text external-service names to a canonical label.

    Extraction produced 51 variants for ~25 real services, mostly differing by a trailing
    parenthetical qualifier ('Gmail (draft only)', 'GitHub (gh CLI)'). Strip the qualifier,
    then map known synonyms.
    """
    raw = (s or '').strip()
    if not raw:
        return None
    base = re.sub(r'\s*[\(（].*$', '', raw).strip(' /·')
    key = base.lower()
    if key in EXT_MAP:
        return EXT_MAP[key]
    for k, v in EXT_MAP.items():
        if key.startswith(k):
            return v
    return base or raw


def ks_key(k):
    """Dedup key for a knowledge source: the Drive ID / URL is authoritative when present,
    otherwise a normalised name (parentheticals and path noise stripped)."""
    idu = (k.get('id_or_url') or '').strip()
    if idu and idu not in ('./references/', '', '-'):
        m = re.search(r'[-\w]{25,}', idu)
        return ('id', (m.group(0) if m else idu).lower())
    nm = re.sub(r'\s*[\(（].*$', '', (k.get('name') or '')).strip()
    nm = re.sub(r'^\./', '', nm).strip().lower()
    return ('name', nm, k.get('type'))


def jl(xs, sep=' · '):
    return sep.join([str(x) for x in xs if x])


def ks_render(ks):
    """Render knowledge sources: '⟳ name (Type | purpose) -> id'  (⟳ = re-read every run)."""
    out = []
    for k in ks:
        flag = '⟳ ' if k.get('runtime_read') else ''
        s = f"{flag}{k.get('name','')} ({KST.get(k.get('type'), k.get('type',''))}"
        if k.get('purpose'):
            s += f" | {k['purpose']}"
        s += ')'
        if k.get('id_or_url'):
            s += f" → {k['id_or_url']}"
        out.append(s)
    return '\n'.join(out)


def style_header(ws, ncols, freeze='D2', row=1, height=42):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill('solid', fgColor=INK_SOFT)
        cell.font = Font(name=FONT_ZH, bold=True, size=10, color=PAPER)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height
    if freeze:
        ws.freeze_panes = freeze


def style_body(ws, ncols, first=2, mono=(), wrap=(), zebra=True):
    for r in range(first, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=(FONT_MONO if c in mono else FONT_ZH), size=9, color=INK)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=(c in wrap))
            cell.border = BORDER
            cell.fill = PatternFill('solid', fgColor=(PAPER_RAIS if (zebra and r % 2 == 0) else PAPER))


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================== Overview
def tab_overview(wb, ctx):
    ws = wb.create_sheet('Overview')
    ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 150])
    rows = [
        ('TITLE', 'Zynkr Skills Index — what we have, what it reads, what it needs', ''),
        ('KV', 'Purpose', 'A human-readable index of every skill published on the Zynkr AI skills marketplace '
                          '(zynkr.ai/ai-skills-marketplace). For each skill it answers four questions: what does it do, '
                          'what sub-skills sit under it, WHERE does it read its knowledge from (which Google Doc / Sheet / '
                          'KB / table), and what does it depend on to run — one of our own products, an MCP server, or an '
                          'external service. Companion to the [3.3] PM Flow and [2.5] Consultant Flow assessments, but this '
                          'one is a portfolio INDEX rather than a coverage assessment: it describes what exists, not what is missing.'),
        ('KV', 'Sources', ctx['sources']),
        ('KV', 'How to read', "Tab 'Skill Index' is the main table — one skill per row, with parent skills bundling their "
                              "└-prefixed children / agents (the shape of the 'PM skills' tab). Tab 'Knowledge Sources' flips "
                              "the same data around to one row per source, so you can see which skills break if you edit a given "
                              "Doc. Tab 'MCP & Services' is the dependency matrix — use it when an MCP server is down or "
                              "unauthorised. Tab 'Solution Map' lists every touchpoint against our own apps. Tab 'Category "
                              "Summary' is the per-category count."),
        ('KV', 'Legend', 'Kind: Orchestrator = calls other skills · Skill = standalone · Sub-skill = only invoked inside its '
                         'parent (shown as a └ row carrying its parent\'s No.). Status: Done / WIP as declared in the skill '
                         'source. Live-read (⟳): TRUE means the knowledge source is fetched fresh on every run, so editing that '
                         'Doc changes behaviour with no prompt edit — the single most important column for maintenance. '
                         'Write impact runs Read-only → Creates drafts → Writes database → Publishes live. MCP required = cannot '
                         'run without that server; in the matrix ✓ = the skill calls that server itself, (✓) = inherited from a '
                         'sub-skill it invokes. Ready to run is the prerequisite gate: Just invoke = no dependencies · Local '
                         'config = a config file or key · Cloud docs = the Drive/DB sources must exist · MCP auth = needs an MCP '
                         'server · MCP + cloud docs = both.'),
        ('GAP', '', ''),
        ('SECTION', 'Headline', ''),
    ]
    for k, v in ctx['headline']:
        rows.append(('KV', k, v))
    rows += [
        ('GAP', '', ''),
        ('KV', 'One-line verdict', ctx['verdict']),
        ('GAP', '', ''),
        ('KV', 'Maintenance', 'Two layers of truth, do not confuse them. Whether a skill is PUBLISHED is decided by the live '
                              'marketplace API (Supabase-backed) — that is the row set. What a skill MEANS (its knowledge '
                              'sources, MCP needs, solution touchpoints) is decided by its SKILL.md in zynkr-skill-builder — '
                              'that is the columns. Re-run the generator after any skill change to rebuild the workbook; do not '
                              'hand-edit cells, they will be overwritten.'),
        ('KV', 'Known caveat', ctx['caveat']),
        ('KV', 'Related', '[3.3] PM Flow × Skill Portfolio Assessment (Sheet 15uxtdPhF95jP4qt0AhdulxuOP5eOwSD9K8yijZ3tqSg) · '
                          '[2.5] Consultant Flow × Skill Portfolio Assessment (Sheet 1Q1o2nYCBJ-uj9hYjcOzua0M4ch_iUMwg_SRdQGvvuyM) · '
                          'taxonomy.md (canonical 0–9 categories) · SKILL_SPEC.md (frontmatter contract) · '
                          'architecture.md (ingest pipeline: SKILL.md → generated JSON → Supabase → marketplace)'),
    ]
    r = 1
    for kind, a, b in rows:
        if kind == 'GAP':
            r += 1
            continue
        if kind == 'TITLE':
            c = ws.cell(row=r, column=1, value=a)
            c.font = Font(name=FONT_ZH, bold=True, size=18, color=INK)
            ws.row_dimensions[r].height = 30
        elif kind == 'SECTION':
            c = ws.cell(row=r, column=1, value=a)
            c.font = Font(name=FONT_ZH, bold=True, size=13, color=SAGE_DEEP)
            ws.row_dimensions[r].height = 26
        else:
            c1 = ws.cell(row=r, column=1, value=a)
            c1.font = Font(name=FONT_ZH, bold=True, size=10, color=INK)
            c1.alignment = Alignment(vertical='top', wrap_text=True)
            c1.fill = PatternFill('solid', fgColor=SAGE_SOFT)
            c1.border = BORDER
            c2 = ws.cell(row=r, column=2, value=b)
            c2.font = Font(name=FONT_ZH, size=10, color=INK_MID)
            c2.alignment = Alignment(vertical='top', wrap_text=True)
            c2.border = BORDER
            ws.row_dimensions[r].height = max(20, 13 * (len(str(b)) // 110 + 1))
        r += 1
    return ws


# ==================================================== Skill Index (main)
def readiness(e):
    """Controlled-vocabulary 'what must be true before I can invoke this' gate.

    The analogue of the reference sheet's Coverage status column: one scannable label instead of
    reading three other columns. Ordered most- to least-demanding so filtering is meaningful.
    """
    mcp = bool(e.get('requires_mcp'))
    setup = [s for s in (e.get('setup_required') or []) if s]
    cloud = any((k.get('type') or '').startswith(('google', 'drive')) or
                (k.get('type') or '') in ('supabase-kb', 'supabase-table')
                for k in (e.get('knowledge_sources') or []))
    if mcp and cloud:
        return 'MCP + cloud docs'
    if mcp:
        return 'MCP auth'
    if cloud:
        return 'Cloud docs'
    if setup:
        return 'Local config'
    return 'Just invoke'


HEADERS = [
    'No.', 'Category', 'Skill / parent', 'Child / agent', 'Kind', 'Status', 'Ready to run',
    'What it does', '一句話說明', 'Triggers', 'Input', 'Process', 'Output',
    'Knowledge source', 'Live-read ⟳', 'Connected solution', 'Solution touchpoint',
    'MCP required', 'MCP servers', 'Key MCP tools', 'External services',
    'Write impact', 'Human gate', 'Artifacts', 'Setup required', 'Caveats',
    'Related skills', 'Provenance', 'Install', 'Source path', 'Updated',
]


def tab_index(wb, inv, sx, ax):
    ws = wb.create_sheet('Skill Index')
    ws.append(HEADERS)
    kids = collections.defaultdict(list)
    for r in inv:
        if r['is_agent'] and r['parent_slug']:
            kids[r['parent_slug']].append(r)
    parents = [r for r in inv if not r['is_agent']]
    parents.sort(key=lambda r: (CATS.get(r['category'], ('9',))[0], r['id']))

    wip, child_rows, parent_rows = [], [], []
    no = 0
    for p in parents:
        no += 1
        e = sx.get(p['id'], {})
        ks = e.get('knowledge_sources') or []
        cs = e.get('connected_solutions') or []
        prov = 'First-party'
        if p.get('upstream_repo') or p.get('original_source_url'):
            prov = f"Adapted — upstream {p.get('upstream_repo') or p.get('original_source_url')}"
        ws.append([
            no, cat_label(p['category']), f"{p['slug']} ({p['id']})", '',
            KIND.get(p['kind'], p['kind']), p.get('status') or '', readiness(e),
            e.get('one_liner_en') or (p.get('summary') or ''), e.get('one_liner_zh', ''),
            jl(e.get('trigger_phrases') or [], '  ·  '),
            p.get('input') or '', p.get('process') or '', p.get('output') or '',
            ks_render(ks), ('TRUE' if any(k.get('runtime_read') for k in ks) else ('FALSE' if ks else '—')),
            jl(sorted({c.get('solution', '') for c in cs})),
            jl([f"{c.get('solution','')}: {c.get('touchpoint','')} ({c.get('direction','')})" for c in cs], '\n'),
            'Yes' if e.get('requires_mcp') else 'No',
            jl(sorted(e.get('mcp_servers') or [])), jl(sorted(e.get('mcp_tools') or [])),
            jl(sorted({x for x in (norm_ext(s) for s in (e.get('external_services') or [])) if x})),
            SIDE.get(e.get('side_effect_level'), e.get('side_effect_level') or ''),
            e.get('human_gate') or '', jl(e.get('artifacts_produced') or []),
            jl(e.get('setup_required') or []), e.get('gotchas') or '',
            jl(sorted(p.get('synergy') or [])), prov,
            p.get('install_command') or '', p.get('source_path') or '', p.get('updated_at') or '',
        ])
        parent_rows.append(ws.max_row)
        if (p.get('status') or '').upper() == 'WIP':
            wip.append(ws.max_row)
        for k in sorted(kids.get(p['slug'], []), key=lambda x: (ax.get(x['id'], {}).get('stage_order') or 99, x['id'])):
            ae = ax.get(k['id'], {})
            order = ae.get('stage_order') or 0
            ws.append([
                no, '', '└', f"{k['slug']} ({k['id']})", 'Sub-skill', k.get('status') or '',
                '(via parent)',
                (f"[stage {order}] " if order else '') + (ae.get('role_en') or k.get('summary') or ''),
                ae.get('role_zh', ''), '',
                ae.get('input') or k.get('input') or '', '', ae.get('output') or k.get('output') or '',
                jl(ae.get('knowledge_sources') or [], '\n'), '', '', '',
                'Yes' if ae.get('requires_mcp') else 'No',
                jl(sorted(ae.get('mcp_servers') or [])), '', '',
                '', '', '', '', '',
                '', 'First-party', '', k.get('source_path') or '', k.get('updated_at') or '',
            ])
            child_rows.append(ws.max_row)
            if (k.get('status') or '').upper() == 'WIP':
                wip.append(ws.max_row)

    n = len(HEADERS)
    style_header(ws, n, freeze='E2', height=46)
    wrapc = {2, 8, 9, 10, 11, 12, 13, 14, 16, 17, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 30}
    style_body(ws, n, mono={1, 3, 4, 29, 30}, wrap=wrapc)
    set_widths(ws, [5, 24, 34, 34, 13, 8, 17, 50, 32, 46, 40, 46, 40,
                    62, 11, 22, 44, 9, 22, 40, 22, 24, 34, 34, 32, 46, 16, 26, 52, 56, 11])
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 58
        ws.cell(row=r, column=1).font = Font(name=FONT_MONO, size=9, bold=True, color=MUTE)
        m = ws.cell(row=r, column=18)
        if m.value == 'Yes':
            m.font = Font(name=FONT_ZH, size=9, bold=True, color=SAGE_DEEP)
        lr = ws.cell(row=r, column=15)
        if lr.value == 'TRUE':
            lr.font = Font(name=FONT_ZH, size=9, bold=True, color=SAGE_DEEP)
        rr = ws.cell(row=r, column=7)
        if rr.value == 'Just invoke':
            rr.font = Font(name=FONT_ZH, size=9, bold=True, color=SAGE_DEEP)
        elif rr.value == '(via parent)':
            rr.font = Font(name=FONT_ZH, size=9, italic=True, color=MUTE)
    for r in parent_rows:
        ws.cell(row=r, column=3).font = Font(name=FONT_MONO, size=10, bold=True, color=INK)
        ws.cell(row=r, column=2).font = Font(name=FONT_ZH, size=9, bold=True, color=INK_MID)
        for c in range(1, n + 1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PAPER_RAIS)
        ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=SAGE_SOFT)
    for r in child_rows:
        ws.cell(row=r, column=3).font = Font(name=FONT_MONO, size=11, bold=True, color=SAGE_DEEP)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=4).font = Font(name=FONT_MONO, size=9, color=INK_MID)
        for c in range(1, n + 1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PAPER)
        ws.row_dimensions[r].height = 40
    for r in wip:
        ws.cell(row=r, column=6).font = Font(name=FONT_ZH, size=9, bold=True, color=ORANGE)
    ws.auto_filter.ref = f'A1:{get_column_letter(n)}{ws.max_row}'
    return ws


# ==================================================== Knowledge Sources
def tab_knowledge(wb, inv, sx):
    ws = wb.create_sheet('Knowledge Sources')
    heads = ['Scope', 'Source', 'Type', 'ID / URL', 'Live-read ⟳', 'Read for', '# skills', 'Consumed by']
    ws.append(heads)
    SCOPE = {
        'google-doc': (1, 'Google Workspace'), 'google-sheet': (1, 'Google Workspace'),
        'google-slides': (1, 'Google Workspace'), 'drive-folder': (1, 'Google Workspace'),
        'supabase-table': (2, 'Database'), 'supabase-kb': (2, 'Database'),
        'gmail': (3, 'Mailbox / calendar'), 'calendar': (3, 'Mailbox / calendar'),
        'repo-file': (4, 'In-repo file'), 'local-file': (4, 'In-repo file'),
        'github': (5, 'Web / GitHub'), 'web': (5, 'Web / GitHub'), 'other': (6, 'Other'),
    }
    byid = {r['id']: r for r in inv}
    agg = {}
    for sid, e in sx.items():
        slug = byid.get(sid, {}).get('slug', sid)
        for k in (e.get('knowledge_sources') or []):
            key = ks_key(k)
            a = agg.setdefault(key, {'names': [], 'ids': set(), 'p': set(), 'rt': False,
                                     's': set(), 'types': set(), 'canon': None})
            if key[0] == 'id' and not a['canon']:
                # keep the ORIGINAL case — Drive IDs are case-sensitive, the dedup key is not
                m = re.search(r'[-\w]{25,}', k.get('id_or_url') or '')
                a['canon'] = m.group(0) if m else None
            a['s'].add(slug)
            a['names'].append((k.get('name') or '').strip())
            a['types'].add(k.get('type', ''))
            if k.get('id_or_url'):
                a['ids'].add(k['id_or_url'].strip())
            if k.get('purpose'):
                a['p'].add(k['purpose'])
            if k.get('runtime_read'):
                a['rt'] = True
    def sortkey(kv):
        a = kv[1]
        pr = min(SCOPE.get(t, (6, 'Other'))[0] for t in a['types'])
        return (pr, -len(a['s']), min(a['names'], key=len) if a['names'] else '')

    band = []
    for key, a in sorted(agg.items(), key=sortkey):
        # display the shortest name seen (longer variants are the same source plus a qualifier)
        name = min(a['names'], key=len) if a['names'] else ''
        pr, label = min((SCOPE.get(t, (6, 'Other')) for t in a['types']), key=lambda x: x[0])
        # When every variant resolves to the same Drive ID, show that ID once, not each phrasing.
        if key[0] == 'id' and a['canon']:
            idcell = a['canon']
        else:
            idcell = jl(sorted(a['ids']), ' | ')
        purposes = sorted(a['p'], key=len)
        pcell = jl(purposes[:5], '; ')
        if len(purposes) > 5:
            pcell += f' (+{len(purposes) - 5} more)'
        ws.append([label, name, jl(sorted(KST.get(t, t) for t in a['types'])), idcell,
                   'TRUE' if a['rt'] else 'FALSE', pcell,
                   len(a['s']), jl(sorted(a['s']))])
        band.append((ws.max_row, pr))
    n = len(heads)
    style_header(ws, n, freeze='C2')
    style_body(ws, n, mono={4}, wrap={2, 4, 6, 8})
    set_widths(ws, [20, 46, 18, 52, 11, 56, 8, 68])
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 38
        if ws.cell(row=r, column=5).value == 'TRUE':
            ws.cell(row=r, column=5).font = Font(name=FONT_ZH, size=9, bold=True, color=SAGE_DEEP)
        ws.cell(row=r, column=2).font = Font(name=FONT_ZH, size=9, bold=True, color=INK)
    for r, pr in band:
        c = ws.cell(row=r, column=1)
        c.font = Font(name=FONT_ZH, size=9, bold=True, color=(INK if pr <= 2 else MUTE))
        if pr <= 2:
            c.fill = PatternFill('solid', fgColor=SAGE_SOFT)
    ws.auto_filter.ref = f'A1:{get_column_letter(n)}{ws.max_row}'
    return ws


# ==================================================== MCP & Services
def tab_mcp(wb, inv, sx):
    ws = wb.create_sheet('MCP & Services')
    def base(s):
        return re.sub(r'\s*\(via sub-skill\)$', '', s).strip()

    servers = sorted({base(s) for e in sx.values() for s in (e.get('mcp_servers') or [])})
    heads = (['No.', 'Skill', 'Category', 'MCP required'] + [f'MCP · {s}' for s in servers]
             + ['# MCP servers', 'External services', 'Write impact'])
    ws.append(heads)
    rows = [r for r in inv if not r['is_agent']]
    rows.sort(key=lambda r: (CATS.get(r['category'], ('9',))[0], r['id']))
    for i, r in enumerate(rows, start=1):
        e = sx.get(r['id'], {})
        raw = e.get('mcp_servers') or []
        direct = {base(s) for s in raw if 'via sub-skill' not in s}
        inher = {base(s) for s in raw if 'via sub-skill' in s}
        exts = sorted({x for x in (norm_ext(s) for s in (e.get('external_services') or [])) if x})
        ws.append([i, f"{r['slug']} ({r['id']})", cat_label(r['category']),
                   'Yes' if e.get('requires_mcp') else 'No'] +
                  [('✓' if s in direct else ('(✓)' if s in inher else '')) for s in servers] +
                  [len(direct | inher), jl(exts, ', '),
                   SIDE.get(e.get('side_effect_level'), e.get('side_effect_level') or '')])
    ws.append([])
    tot = ['', 'Total', '', sum(1 for r in rows if sx.get(r['id'], {}).get('requires_mcp'))]
    tot += [sum(1 for r in rows if s in {base(x) for x in (sx.get(r['id'], {}).get('mcp_servers') or [])})
            for s in servers]
    tot += ['', '', '']
    ws.append(tot)
    n = len(heads)
    style_header(ws, n, freeze='C2', height=100)
    style_body(ws, n, mono={1, 2}, wrap={3, len(heads) - 1})
    set_widths(ws, [5, 34, 24, 12] + [14] * len(servers) + [12, 54, 26])
    for c in range(5, 5 + len(servers)):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal='left', vertical='bottom',
                                                      wrap_text=True, textRotation=60)
    last = ws.max_row
    for c in range(1, n + 1):
        cell = ws.cell(row=last, column=c)
        cell.font = Font(name=FONT_ZH, size=10, bold=True, color=PAPER)
        cell.fill = PatternFill('solid', fgColor=INK_MID)
    for r in range(2, last):
        ws.row_dimensions[r].height = 24
        ws.cell(row=r, column=1).font = Font(name=FONT_MONO, size=9, color=MUTE)
        ws.cell(row=r, column=2).font = Font(name=FONT_MONO, size=9, bold=True, color=INK)
        for c in range(4, 5 + len(servers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == '✓':
                cell.font = Font(name=FONT_ZH, size=11, bold=True, color=SAGE_DEEP)
            elif cell.value == '(✓)':
                cell.font = Font(name=FONT_ZH, size=9, italic=True, color=MUTE)
    return ws


# ==================================================== Solution Map
def tab_solutions(wb, inv, sx):
    ws = wb.create_sheet('Solution Map')
    heads = ['Solution', 'Skill', 'Category', 'Direction', 'Touchpoint', 'Write impact', 'MCP required']
    ws.append(heads)
    byid = {r['id']: r for r in inv}
    recs = []
    for sid, e in sx.items():
        r = byid.get(sid)
        if not r:
            continue
        for c in (e.get('connected_solutions') or []):
            recs.append((c.get('solution', ''), f"{r['slug']} ({sid})", cat_label(r['category']),
                         c.get('direction', ''), c.get('touchpoint', ''),
                         SIDE.get(e.get('side_effect_level'), e.get('side_effect_level') or ''),
                         'Yes' if e.get('requires_mcp') else 'No'))
    for t in sorted(recs, key=lambda t: (t[0], t[1])):
        ws.append(list(t))
    n = len(heads)
    style_header(ws, n, freeze='B2')
    style_body(ws, n, mono={2}, wrap={3, 5})
    set_widths(ws, [28, 34, 24, 12, 66, 26, 12])
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 32
        ws.cell(row=r, column=1).font = Font(name=FONT_ZH, size=9, bold=True, color=INK)
        ws.cell(row=r, column=2).font = Font(name=FONT_MONO, size=9, color=INK)
    ws.auto_filter.ref = f'A1:{get_column_letter(n)}{ws.max_row}'
    return ws


# ==================================================== Category Summary
def tab_summary(wb, inv, sx):
    ws = wb.create_sheet('Category Summary')
    heads = ['#', 'Category', 'Skills', 'Sub-skills', 'Done', 'WIP', 'Needs MCP',
             'Reads knowledge', 'Live-read ⟳', 'Connected to our apps', 'Writes or publishes', 'Examples']
    ws.append(heads)
    skills = [r for r in inv if not r['is_agent']]
    agents = [r for r in inv if r['is_agent']]
    for slug, (num, en, zh) in sorted(CATS.items(), key=lambda kv: kv[1][0]):
        srs = [r for r in skills if r['category'] == slug]
        ars = [r for r in agents if r['category'] == slug]
        if not srs and not ars:
            ws.append([num, f'{num}. {en} {zh}', 0, 0, 0, 0, 0, 0, 0, 0, 0, '(no skills yet)'])
            continue
        exs = [sx.get(r['id'], {}) for r in srs]
        ws.append([
            num, f'{num}. {en} {zh}', len(srs), len(ars),
            sum(1 for r in srs if (r.get('status') or '') == 'Done'),
            sum(1 for r in srs if (r.get('status') or '') == 'WIP'),
            sum(1 for e in exs if e.get('requires_mcp')),
            sum(1 for e in exs if e.get('knowledge_sources')),
            sum(1 for e in exs if any(k.get('runtime_read') for k in (e.get('knowledge_sources') or []))),
            sum(1 for e in exs if e.get('connected_solutions')),
            sum(1 for e in exs if e.get('side_effect_level') in ('writes-database', 'publishes-live', 'mixed')),
            jl([r['slug'] for r in sorted(srs, key=lambda x: x['id'])[:3]]),
        ])
    exs = [sx.get(r['id'], {}) for r in skills]
    ws.append([])
    ws.append(['', 'Total', len(skills), len(agents),
               sum(1 for r in skills if (r.get('status') or '') == 'Done'),
               sum(1 for r in skills if (r.get('status') or '') == 'WIP'),
               sum(1 for e in exs if e.get('requires_mcp')),
               sum(1 for e in exs if e.get('knowledge_sources')),
               sum(1 for e in exs if any(k.get('runtime_read') for k in (e.get('knowledge_sources') or []))),
               sum(1 for e in exs if e.get('connected_solutions')),
               sum(1 for e in exs if e.get('side_effect_level') in ('writes-database', 'publishes-live', 'mixed')), ''])
    n = len(heads)
    style_header(ws, n, freeze='C2')
    style_body(ws, n, mono={1}, wrap={2, 12})
    set_widths(ws, [5, 32, 8, 11, 7, 7, 11, 15, 12, 20, 18, 52])
    last = ws.max_row
    for c in range(1, n + 1):
        cell = ws.cell(row=last, column=c)
        cell.font = Font(name=FONT_ZH, size=10, bold=True, color=PAPER)
        cell.fill = PatternFill('solid', fgColor=INK_MID)
    for r in range(2, last):
        ws.row_dimensions[r].height = 24
        ws.cell(row=r, column=2).font = Font(name=FONT_ZH, size=10, bold=True, color=INK)
        for c in range(3, 12):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
    return ws


def rollup(inv, sx, ax):
    """A parent orchestrator inherits its children's dependencies.

    The audit found parents (e.g. zynkr-content-writer 1.01) whose only MCP / Drive evidence
    lives in agents/*.md one level down. For an index that reads as "this skill needs no MCP",
    which is wrong at runtime — invoking the parent invokes the children. Union the child
    dependencies into the parent and tag the inherited-only ones so the provenance stays visible.
    """
    byslug = {r['slug']: r for r in inv}
    kids = collections.defaultdict(list)
    for r in inv:
        if r['is_agent'] and r['parent_slug']:
            kids[r['parent_slug']].append(r)
    n_fixed = 0
    for pslug, ks in kids.items():
        p = byslug.get(pslug)
        if not p or p['id'] not in sx:
            continue
        e = sx[p['id']]
        own = set(e.get('mcp_servers') or [])
        inherited = set()
        for k in ks:
            inherited |= set((ax.get(k['id'], {}) or {}).get('mcp_servers') or [])
        extra = inherited - own
        if extra:
            e['mcp_servers'] = sorted(own) + [f'{s} (via sub-skill)' for s in sorted(extra)]
            n_fixed += 1
        if (extra or inherited) and not e.get('requires_mcp'):
            e['requires_mcp'] = True
            e['_mcp_inherited'] = True
        ks_names = {(k.get('name') or '').strip() for k in (e.get('knowledge_sources') or [])}
        for k in ks:
            for nm in ((ax.get(k['id'], {}) or {}).get('knowledge_sources') or []):
                nm = (nm or '').strip()
                if nm and nm not in ks_names:
                    ks_names.add(nm)
                    e.setdefault('knowledge_sources', []).append({
                        'type': 'other', 'name': nm, 'id_or_url': '', 'runtime_read': False,
                        'purpose': f"via sub-skill {k['slug']}",
                    })
    return n_fixed


def main():
    inv = json.load(open(J + '/inventory.json'))
    ex = json.load(open(J + '/extracted.json'))
    sx = {r['id']: r for r in ex.get('skills', [])}
    ax = {r['id']: r for r in ex.get('subagents', [])}
    n_roll = rollup(inv, sx, ax)
    print(f'rollup: {n_roll} parent skills inherited MCP servers from their sub-skills')
    skills = [r for r in inv if not r['is_agent']]
    agents = [r for r in inv if r['is_agent']]
    exs = [sx.get(r['id'], {}) for r in skills]
    today = datetime.date.today().isoformat()

    n_mcp = sum(1 for e in exs if e.get('requires_mcp'))
    n_kn = sum(1 for e in exs if e.get('knowledge_sources'))
    n_rt = sum(1 for e in exs if any(k.get('runtime_read') for k in (e.get('knowledge_sources') or [])))
    n_sol = sum(1 for e in exs if e.get('connected_solutions'))
    n_risk = sum(1 for e in exs if e.get('side_effect_level') in ('writes-database', 'publishes-live', 'mixed'))
    n_pub = sum(1 for e in exs if e.get('side_effect_level') in ('writes-database', 'publishes-live'))
    n_ro = sum(1 for e in exs if e.get('side_effect_level') == 'read-only')
    n_orc = sum(1 for r in skills if r['kind'] == 'orchestrator')
    n_par = len({r['parent_slug'] for r in agents if r['parent_slug']})
    srv = collections.Counter(s for e in exs for s in (e.get('mcp_servers') or []))
    ext = collections.Counter(s for e in exs for s in (e.get('external_services') or []))
    cats_used = len({r['category'] for r in skills})
    missing = [r['id'] for r in skills if r['id'] not in sx]
    n_gate = sum(1 for e in exs if (e.get('human_gate') or '').strip().lower() in ('', 'none'))
    order = ['Just invoke', 'Local config', 'Cloud docs', 'MCP auth', 'MCP + cloud docs']
    rc = collections.Counter(readiness(e) for e in exs)
    readymix = [(k, rc[k]) for k in order if rc.get(k)]

    ctx = {
        'sources': f'Row set: live marketplace API zynkr.ai/api/skills — {len(inv)} published entries, snapshot {today}. '
                   f'Column semantics: the SKILL.md source files in peter-tu-zynkr/zynkr-skill-builder '
                   f'(skills/<0-9>-<category>/<slug>/SKILL.md plus agents/*.md), read in full. '
                   f'Categories: taxonomy.md (canonical 0–9). Design tokens: zynkr-website-fe/styles.css :root.',
        'headline': [
            ('Entries indexed', f'{len(inv)} — {len(skills)} callable skills + {len(agents)} sub-skills'),
            ('Of the callable skills', f'{n_orc} orchestrators (call other skills) · {len(skills) - n_orc} standalone'),
            ('Sub-skills', f'{len(agents)} across {n_par} parent skills; the rest are single-agent skills'),
            ('Categories in use', f'{cats_used} of 10 taxonomy categories carry skills'),
            ('Status', f"Done {sum(1 for r in skills if r.get('status') == 'Done')} · "
                       f"WIP {sum(1 for r in skills if r.get('status') == 'WIP')}"),
            ('Reads external knowledge', f'{n_kn} of {len(skills)} skills ({round(100*n_kn/max(1,len(skills)))}%) read a Doc / Sheet / KB / table beyond user input'),
            ('Live-read ⟳', f'{n_rt} skills re-read their source every run — editing those documents changes behaviour with no code change'),
            ('Needs an MCP server', f'{n_mcp} of {len(skills)} skills ({round(100*n_mcp/max(1,len(skills)))}%). '
                                    f'Most-depended: ' + jl([f'{k} ({v})' for k, v in srv.most_common(6)])),
            ('External services', jl([f'{k} ({v})' for k, v in ext.most_common(8)]) or 'none'),
            ('Connected to our own apps', f'{n_sol} skills touch a Zynkr product (AI 平台 / CMS / Accounting / Learn / website / admin)'),
            ('Write impact', f'{n_ro} skills are read-only. {n_risk} write a database, publish live, or mix side effects '
                             f'({n_pub} do so unambiguously) — those are the ones worth a human gate, and '
                             f'{len(skills) - n_gate} of the {len(skills)} declare one'),
            ('Ready to run', jl([f'{k} ({v})' for k, v in readymix])),
        ],
        'verdict': f'The portfolio is knowledge-grounded and platform-connected rather than self-contained: '
                   f'{n_kn} of {len(skills)} skills read their standards from somewhere else, and {n_rt} of those re-read on every '
                   f'run — which means a large share of behaviour is tuned by editing Google Docs, not prompts. '
                   f'{n_mcp} skills cannot run without an MCP server, so MCP auth is a single point of failure across the fleet; '
                   f'the "MCP & Services" tab is the blast-radius map for that. The heaviest concentration is Brand & Marketing '
                   f'and Sales & Consultant, which together hold the majority of skills and nearly all of the live-publishing ones.',
        'caveat': 'Columns are derived by reading each SKILL.md (and its agents/ + references/ files) in full, so they report what '
                  'the source declares. Treat an empty Knowledge source cell as "nothing declared in the source", not a guarantee '
                  'of none. Three specifics worth knowing: (1) an MCP server tagged "(via sub-skill)" is inherited — the parent '
                  'itself makes no such call, but invoking it invokes a child that does, so the dependency is real at runtime; '
                  '(2) consult-flow-design (2.14) names Lucid throughout but makes no Lucid call itself — it is a thin delegation '
                  'wrapper over product-flow-design, so its Lucid dependency is transitive; (3) zynkr-gm (0.02) is published but '
                  'has no file in the repo (skills/0-strategy/ holds only a .gitkeep) — its row was read from the published '
                  'zynkr.ai/s/0.02.md, and that source_path is the one broken pointer of the 112. Also unindexed: a nested child '
                  'skill at zynkr-content-writer/.claude/skills/write-article/SKILL.md, which the marketplace does not list.'
                  + (f' Extraction gap: {len(missing)} skill(s) had no metadata extracted ({jl(missing)}).' if missing else ''),
    }

    wb = Workbook()
    wb.remove(wb.active)
    tab_overview(wb, ctx)
    tab_index(wb, inv, sx, ax)
    tab_knowledge(wb, inv, sx)
    tab_mcp(wb, inv, sx)
    tab_solutions(wb, inv, sx)
    tab_summary(wb, inv, sx)
    wb.save(OUT)
    print('WROTE', OUT)
    for s in wb.sheetnames:
        print(f'  {s:22s} {wb[s].max_row:4d} rows x {wb[s].max_column:3d} cols')
    print('extracted skills:', len(sx), '/', len(skills), '| subagents:', len(ax), '/', len(agents))
    if missing:
        print('MISSING extraction:', missing)


if __name__ == '__main__':
    main()
